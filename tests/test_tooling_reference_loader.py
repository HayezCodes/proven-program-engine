"""
test_tooling_reference_loader.py — Tests for the tooling_reference_loader module.
"""

import pytest
import pandas as pd
from pathlib import Path

from src.tooling_reference_loader import (
    load_tooling_reference,
    export_tooling_reference,
    _extract_machine_id,
    _safe_tool_number,
    _normalize_description,
)


# ---------------------------------------------------------------------------
# _extract_machine_id
# ---------------------------------------------------------------------------

class TestExtractMachineId:
    def test_plain_number(self):
        assert _extract_machine_id("655") == "655"

    def test_number_with_name(self):
        assert _extract_machine_id("655 Haas Endmill Sizes 2/12/2026") == "655"

    def test_four_digit(self):
        assert _extract_machine_id("1234 Machine") == "1234"

    def test_leading_spaces(self):
        assert _extract_machine_id("  432  MAZAK LATHE") == "432"

    def test_no_number_returns_none(self):
        assert _extract_machine_id("Haas Machine") is None

    def test_empty_string_returns_none(self):
        assert _extract_machine_id("") is None


# ---------------------------------------------------------------------------
# _safe_tool_number
# ---------------------------------------------------------------------------

class TestSafeToolNumber:
    def test_integer_value(self):
        assert _safe_tool_number(1) == 1

    def test_float_whole_number(self):
        assert _safe_tool_number(17.0) == 17

    def test_fractional_float_returns_none(self):
        assert _safe_tool_number(1.5) is None

    def test_string_integer(self):
        assert _safe_tool_number("22") == 22

    def test_range_text_returns_none(self):
        assert _safe_tool_number("10.6/11.6") is None

    def test_none_input(self):
        assert _safe_tool_number(None) is None

    def test_nan_input(self):
        assert _safe_tool_number(float("nan")) is None

    def test_zero_returns_none(self):
        assert _safe_tool_number(0) is None


# ---------------------------------------------------------------------------
# _normalize_description
# ---------------------------------------------------------------------------

class TestNormalizeDescription:
    def test_strips_whitespace(self):
        assert _normalize_description("  1/2  ") == "1/2"

    def test_returns_none_for_none(self):
        assert _normalize_description(None) is None

    def test_returns_none_for_blank_string(self):
        assert _normalize_description("   ") is None

    def test_preserves_content(self):
        assert _normalize_description("WOODRUFF ") == "WOODRUFF"

    def test_mixed_case_preserved(self):
        assert _normalize_description("120° SPOT DRILL ") == "120° SPOT DRILL"


# ---------------------------------------------------------------------------
# Mock workbook fixture
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_wb_path(tmp_path):
    """Write a minimal mock Milling Tools workbook matching the real sheet structure."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milling Tools"

    # Row 1 (openpyxl row 1 = pandas row 0): empty
    # Row 2 (openpyxl row 2 = pandas row 1): section headers
    ws.cell(row=2, column=2, value="655 Haas Endmill Sizes  2/12/2026")  # pandas col 1
    ws.cell(row=2, column=9, value="432  MAZAK LATHE Endmill Sizes")     # pandas col 8
    ws.cell(row=2, column=13, value="654 OKUMA Tool List")               # pandas col 12

    # Row 3 (openpyxl row 3 = pandas row 2): column headers
    ws.cell(row=3, column=2, value="TOOL# ")
    ws.cell(row=3, column=3, value="Fraction")
    ws.cell(row=3, column=4, value="Decimal")
    ws.cell(row=3, column=9, value="TOOL# ")
    ws.cell(row=3, column=10, value="Fraction")
    ws.cell(row=3, column=11, value="Decimal")
    ws.cell(row=3, column=13, value="#")
    ws.cell(row=3, column=14, value="Fraction")
    ws.cell(row=3, column=15, value="Decimal")

    # Row 4 (openpyxl row 4 = pandas row 3): first data row
    # 655 Haas T1=1/2
    ws.cell(row=4, column=2, value=1)
    ws.cell(row=4, column=3, value="1/2")
    ws.cell(row=4, column=4, value=0.5)
    # 432 Mazak: range text + description
    ws.cell(row=4, column=9, value="10.6/11.6")
    ws.cell(row=4, column=10, value="1/2")
    ws.cell(row=4, column=11, value=0.5)
    # 654 Okuma T1=1/2
    ws.cell(row=4, column=13, value=1)
    ws.cell(row=4, column=14, value="1/2")
    ws.cell(row=4, column=15, value=0.5)

    # Row 5 (pandas row 4): T2
    ws.cell(row=5, column=2, value=2)
    ws.cell(row=5, column=3, value="12mm")
    ws.cell(row=5, column=4, value=0.4724)
    ws.cell(row=5, column=10, value="3/8")
    ws.cell(row=5, column=11, value=0.375)
    ws.cell(row=5, column=13, value=2)
    ws.cell(row=5, column=14, value="12mm")
    ws.cell(row=5, column=15, value=0.4724)

    # Row 6 (pandas row 5): T3 — Haas T3 with no description (empty slot)
    ws.cell(row=6, column=2, value=3)
    ws.cell(row=6, column=13, value=3)
    ws.cell(row=6, column=14, value="WOODRUFF ")

    out = tmp_path / "test_tooling_wb.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# load_tooling_reference
# ---------------------------------------------------------------------------

class TestLoadToolingReference:
    def test_returns_list(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        assert isinstance(result, list)

    def test_has_records(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        assert len(result) > 0

    def test_haas_machine_id(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        haas = [r for r in result if r["machine_id"] == "655"]
        assert len(haas) > 0

    def test_okuma_machine_id(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        okuma = [r for r in result if r["machine_id"] == "654"]
        assert len(okuma) > 0

    def test_mazak_machine_id(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        mazak = [r for r in result if r["machine_id"] == "432"]
        assert len(mazak) > 0

    def test_haas_t1_description(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        t1 = next(r for r in result if r["machine_id"] == "655" and r["tool_number"] == 1)
        assert t1["description"] == "1/2"

    def test_haas_t1_decimal(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        t1 = next(r for r in result if r["machine_id"] == "655" and r["tool_number"] == 1)
        assert t1["decimal_size"] == 0.5

    def test_okuma_t1_description(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        t1 = next(r for r in result if r["machine_id"] == "654" and r["tool_number"] == 1)
        assert t1["description"] == "1/2"

    def test_okuma_woodruff_description(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        woodruff = next(r for r in result if r["machine_id"] == "654" and r["tool_number"] == 3)
        assert woodruff["description"] == "WOODRUFF"

    def test_haas_empty_slot_needs_review(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        t3 = next((r for r in result if r["machine_id"] == "655" and r["tool_number"] == 3), None)
        assert t3 is not None
        assert t3["needs_review"] is True

    def test_haas_t1_not_needs_review(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        t1 = next(r for r in result if r["machine_id"] == "655" and r["tool_number"] == 1)
        assert t1["needs_review"] is False

    def test_mazak_needs_review(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        mazak = [r for r in result if r["machine_id"] == "432"]
        assert all(r["needs_review"] is True for r in mazak)

    def test_mazak_inferred_tool_number(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        mazak = [r for r in result if r["machine_id"] == "432"]
        # First Mazak entry should have tool_number=1 (inferred from row position)
        assert mazak[0]["tool_number"] == 1

    def test_mazak_range_text_preserved_in_raw(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        mazak = [r for r in result if r["machine_id"] == "432"]
        assert mazak[0]["tool_number_raw"] == "10.6/11.6"

    def test_record_has_required_keys(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        required = {
            "machine_id", "machine_label", "tool_number", "tool_number_raw",
            "description", "decimal_size", "needs_review",
        }
        assert required.issubset(result[0].keys())

    def test_machine_label_is_string(self, mock_wb_path):
        result = load_tooling_reference(mock_wb_path)
        assert all(isinstance(r["machine_label"], str) for r in result)


# ---------------------------------------------------------------------------
# export_tooling_reference
# ---------------------------------------------------------------------------

MOCK_RECORDS = [
    {
        "machine_id": "655",
        "machine_label": "655 Haas Endmill Sizes",
        "tool_number": 1,
        "tool_number_raw": "1",
        "description": "1/2",
        "decimal_size": 0.5,
        "needs_review": False,
    },
    {
        "machine_id": "655",
        "machine_label": "655 Haas Endmill Sizes",
        "tool_number": 17,
        "tool_number_raw": "17",
        "description": "WOODRUFF",
        "decimal_size": None,
        "needs_review": False,
    },
    {
        "machine_id": "432",
        "machine_label": "432  MAZAK LATHE Endmill Sizes",
        "tool_number": 1,
        "tool_number_raw": "10.6/11.6",
        "description": "1/2",
        "decimal_size": 0.5,
        "needs_review": True,
    },
]


class TestExportToolingReference:
    def test_creates_csv(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        assert path.exists()
        assert path.suffix == ".csv"

    def test_filename_prefix(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        assert path.name.startswith("machine_tooling_reference_")

    def test_timestamp_in_filename(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        assert "20260101_120000" in path.name

    def test_csv_row_count(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert len(df) == len(MOCK_RECORDS)

    def test_csv_has_machine_id_column(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert "machine_id" in df.columns

    def test_csv_has_tool_number_column(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert "tool_number" in df.columns

    def test_csv_has_needs_review_column(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert "needs_review" in df.columns

    def test_auto_timestamp_creates_file(self, tmp_path):
        path = export_tooling_reference(MOCK_RECORDS, tmp_path)
        assert path.exists()
