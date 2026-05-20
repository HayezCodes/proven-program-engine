"""
test_reference_loader.py — Tests for the reference_loader module.
"""

import pytest
import pandas as pd
from pathlib import Path

from src.reference_loader import (
    load_reference, export_reference, _normalize_material, _parse_finish_feed,
)


# ---------------------------------------------------------------------------
# _normalize_material
# ---------------------------------------------------------------------------

class TestNormalizeMaterial:
    def test_strips_leading_trailing_whitespace(self):
        assert _normalize_material("  1018  ") == "1018"

    def test_uppercases(self):
        assert _normalize_material("hastelloy c276") == "HASTELLOY C276"

    def test_collapses_internal_spaces(self):
        assert _normalize_material("S31803  DUPLEX") == "S31803 DUPLEX"

    def test_plain_number_string(self):
        assert _normalize_material("4140") == "4140"

    def test_already_normalized(self):
        assert _normalize_material("TITANIUM #2") == "TITANIUM #2"


# ---------------------------------------------------------------------------
# Shared mock reference data
# ---------------------------------------------------------------------------

MOCK_REFERENCE = [
    {
        "material_raw": "1018",
        "material": "1018",
        "turning_rough_sfm": 615.0,
        "turning_finish_sfm": 720.0,
        "turning_rough_doc": 0.100,
        "turning_rough_ipr": 0.018,
        "turning_finish_ipr": 0.016,
        "milling_sfm": 310.0,
        "milling_rough_ipm": 15.0,
        "milling_finish_ipm": 25.0,
    },
    {
        "material_raw": "4140",
        "material": "4140",
        "turning_rough_sfm": 450.0,
        "turning_finish_sfm": 600.0,
        "turning_rough_doc": 0.100,
        "turning_rough_ipr": 0.016,
        "turning_finish_ipr": 0.016,
        "milling_sfm": 310.0,
        "milling_rough_ipm": 10.0,
        "milling_finish_ipm": 16.0,
    },
    {
        "material_raw": "HASTELLOY C276",
        "material": "HASTELLOY C276",
        "turning_rough_sfm": 150.0,
        "turning_finish_sfm": 300.0,
        "turning_rough_doc": None,
        "turning_rough_ipr": 0.010,
        "turning_finish_ipr": 0.012,
        "milling_sfm": 200.0,
        "milling_rough_ipm": 10.0,
        "milling_finish_ipm": 16.0,
    },
]


# ---------------------------------------------------------------------------
# Mock workbook fixture
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_wb_path(tmp_path):
    """Write a minimal mock Excel workbook matching the real sheet structure."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPEEDS AND FEEDS "

    # Row 1: empty
    # Row 2: section labels — match real file (col C = "TURNING ")
    ws.cell(row=2, column=4, value="TURNING ")
    # Row 3: column headers — match real column positions
    # col A(1)=empty, col B(2)=ref, col C(3)=material, col D(4)=rough SFM,
    # col E(5)=finish SFM, col F(6)=DOC, col G(7)=rough IPR, col H(8)=finish IPR,
    # col I(9)=separator, col J(10)=milling SFM, col K(11)=rough IPM, col L(12)=finish IPM
    ws.cell(row=3, column=2, value="ref")
    ws.cell(row=3, column=3, value="STEEL ")
    ws.cell(row=3, column=4, value="ROUGH SPEED ")
    ws.cell(row=3, column=5, value="FINISH SPEED ")
    ws.cell(row=3, column=6, value="ROUGH DOC. ")
    ws.cell(row=3, column=7, value="ROUGH FEED ")
    ws.cell(row=3, column=8, value="FINISH FEED ")
    ws.cell(row=3, column=10, value="milling sfm ")
    ws.cell(row=3, column=11, value="ROUGH FEED")
    ws.cell(row=3, column=12, value="FINISH")
    # Data row 1 (sheet row 4) — ref=1, material="1018"
    ws.cell(row=4, column=2, value=1)
    ws.cell(row=4, column=3, value="1018")
    ws.cell(row=4, column=4, value=615)
    ws.cell(row=4, column=5, value=720)
    ws.cell(row=4, column=6, value=0.100)
    ws.cell(row=4, column=7, value=0.018)
    ws.cell(row=4, column=8, value=0.016)
    ws.cell(row=4, column=10, value=310)
    ws.cell(row=4, column=11, value=15)
    ws.cell(row=4, column=12, value=25)
    # Data row 2 (sheet row 5) — ref=2, material="4140"
    ws.cell(row=5, column=2, value=2)
    ws.cell(row=5, column=3, value="4140")
    ws.cell(row=5, column=4, value=450)
    ws.cell(row=5, column=5, value=600)
    ws.cell(row=5, column=6, value=0.100)
    ws.cell(row=5, column=7, value=0.016)
    ws.cell(row=5, column=8, value=0.016)
    ws.cell(row=5, column=10, value=310)
    ws.cell(row=5, column=11, value=10)
    ws.cell(row=5, column=12, value=16)

    out = tmp_path / "test_wb.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# load_reference
# ---------------------------------------------------------------------------

class TestLoadReference:
    def test_returns_list(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert isinstance(result, list)

    def test_correct_row_count(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert len(result) == 2

    def test_material_normalized(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["material"] == "1018"

    def test_material_raw_preserved(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["material_raw"] == "1018"

    def test_turning_sfm_values(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["turning_rough_sfm"] == 615.0
        assert result[0]["turning_finish_sfm"] == 720.0

    def test_ipr_values(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["turning_rough_ipr"] == 0.018
        assert result[0]["turning_finish_ipr"] == 0.016

    def test_milling_sfm(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["milling_sfm"] == 310.0

    def test_milling_ipm_values(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[0]["milling_rough_ipm"] == 15.0
        assert result[0]["milling_finish_ipm"] == 25.0

    def test_second_material_different_sfm(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert result[1]["material"] == "4140"
        assert result[1]["turning_rough_sfm"] == 450.0


# ---------------------------------------------------------------------------
# export_reference
# ---------------------------------------------------------------------------

class TestExportReference:
    def test_creates_csv(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        assert path.exists()
        assert path.suffix == ".csv"

    def test_timestamp_in_filename(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        assert "20260101_120000" in path.name

    def test_filename_prefix(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        assert path.name.startswith("shop_sf_reference_")

    def test_csv_row_count(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert len(df) == len(MOCK_REFERENCE)

    def test_csv_has_material_column(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert "material" in df.columns

    def test_csv_has_sfm_columns(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path, timestamp="20260101_120000")
        df = pd.read_csv(path)
        assert "turning_rough_sfm" in df.columns
        assert "turning_finish_sfm" in df.columns

    def test_auto_timestamp_creates_file(self, tmp_path):
        path = export_reference(MOCK_REFERENCE, tmp_path)
        assert path.exists()


# ---------------------------------------------------------------------------
# _parse_finish_feed
# ---------------------------------------------------------------------------

class TestParseFinishFeed:
    def test_single_float_sets_finish_feed_mid(self):
        result = _parse_finish_feed(0.016)
        assert result["finish_feed_mid"] == 0.016
        assert result["finish_feed_low"] is None
        assert result["finish_feed_high"] is None

    def test_single_string_value(self):
        result = _parse_finish_feed("0.016")
        assert result["finish_feed_mid"] == 0.016

    def test_two_values_slash_separated(self):
        result = _parse_finish_feed("0.010/0.016")
        assert result["finish_feed_low"] == 0.010
        assert result["finish_feed_high"] == 0.016
        assert result["finish_feed_mid"] is None

    def test_two_values_comma_separated(self):
        result = _parse_finish_feed("0.010, 0.014")
        assert result["finish_feed_low"] == 0.010
        assert result["finish_feed_high"] == 0.014

    def test_three_values_sets_all_bands(self):
        result = _parse_finish_feed("0.008, 0.012, 0.016")
        assert result["finish_feed_low"] == 0.008
        assert result["finish_feed_mid"] == 0.012
        assert result["finish_feed_high"] == 0.016

    def test_raw_text_preserved_for_multi_value(self):
        result = _parse_finish_feed("0.010/0.016")
        assert result["finish_feed_raw"] == "0.010/0.016"

    def test_raw_text_preserved_for_float(self):
        result = _parse_finish_feed(0.016)
        assert "0.016" in result["finish_feed_raw"]

    def test_none_input_returns_empty(self):
        result = _parse_finish_feed(None)
        assert result["finish_feed_mid"] is None
        assert result["finish_feed_raw"] == ""

    def test_three_values_sorted_correctly(self):
        # Provide out-of-order values to confirm sorting
        result = _parse_finish_feed("0.016/0.008/0.012")
        assert result["finish_feed_low"] == 0.008
        assert result["finish_feed_mid"] == 0.012
        assert result["finish_feed_high"] == 0.016


# ---------------------------------------------------------------------------
# finish_feed_* fields in load_reference output
# ---------------------------------------------------------------------------

class TestLoadReferenceFinishFeedFields:
    def test_reference_has_finish_feed_fields(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert "finish_feed_raw" in result[0]
        assert "finish_feed_low" in result[0]
        assert "finish_feed_mid" in result[0]
        assert "finish_feed_high" in result[0]

    def test_single_float_cell_maps_to_finish_feed_mid(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        r_1018 = next(r for r in result if r["material"] == "1018")
        assert r_1018["finish_feed_mid"] == 0.016
        assert r_1018["finish_feed_low"] is None
        assert r_1018["finish_feed_high"] is None

    def test_finish_feed_raw_is_string(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        assert isinstance(result[0]["finish_feed_raw"], str)

    def test_turning_finish_ipr_backward_compat(self, mock_wb_path):
        result = load_reference(mock_wb_path)
        r_1018 = next(r for r in result if r["material"] == "1018")
        # turning_finish_ipr must still be present for backward compatibility
        assert r_1018["turning_finish_ipr"] == 0.016
